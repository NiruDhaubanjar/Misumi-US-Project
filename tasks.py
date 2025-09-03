from celery import Celery
import os
from openpyxl import load_workbook
import pandas as pd
import glob
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Border, Side
from spec_tables import extract_all_tables
from bs4 import BeautifulSoup
import pandas as pd
from basic_information import extract_title, extract_subtitle, extract_part_number, extract_description, extract_volume_discount_table
from specification_alterations import extract_spec_table_from_html, extract_alteration_specs

broker_url = os.getenv("CELERY_BROKER_URL", "redis://redis:6379/0")
backend_url = os.getenv("CELERY_RESULT_BACKEND", "redis://redis:6379/0")

celery_app = Celery("parser", broker=broker_url, backend=backend_url)
celery_app.conf.broker_connection_retry_on_startup = True

celery_app.conf.broker_connection_retry_on_startup = True

# Parser logic merged here
def process_multiple_html_files(html_files):
    all_data = []

    for html_file in html_files:
        with open(html_file, "r", encoding="utf-8") as f:
            html_content = f.read()
            soup = BeautifulSoup(html_content, "html.parser")

        title = extract_title(soup).get("value", "") or ""
        subtitle = extract_subtitle(soup).get("value", "") or ""
        part_number = extract_part_number(soup).get("value", "") or ""
        description = extract_description(soup).get("value", "") or ""

        spec_data = extract_spec_table_from_html(html_file)

        alteration_df = extract_alteration_specs(html_file)
        alteration_data = []
        if not alteration_df.empty:
            for _, row in alteration_df.iterrows():
                alteration_data.extend([row["Spec Name"], row["Value"]])

        volume_data = extract_volume_discount_table(soup)
        volume_flat = []
        for i, row in enumerate(volume_data, start=1):
            volume_flat.extend([
                f"Quantity{i}", row.get("Quantity", ""),
                f"Price{i}", row.get("Price", ""),
                f"ShipDate{i}", row.get("Ship Date", "")
            ])

        row = [
            "Title", title,
            "Subtitle", subtitle,
            "Part Number", part_number,
            "Description", description
        ]
        row += spec_data
        row += volume_flat
        if alteration_data:
            row += ["Alteration"]  
            row += alteration_data

        all_data.append(row)

    max_cols = max(len(row) for row in all_data)
    all_data = [row + [""]*(max_cols - len(row)) for row in all_data]

    return pd.DataFrame(all_data)

# Celery task
@celery_app.task
def process_folder_task(root, input_folder, output_folder):
    """Process all HTMLs in one folder -> save Excel file"""
    html_files = [os.path.join(root, f) for f in os.listdir(root) if f.lower().endswith(".html")]

    if not html_files:
        return f"⚠️ No HTML files in {root}"

    merged_df = process_multiple_html_files(html_files)

    relative_path = os.path.relpath(root, input_folder)
    folder_name = os.path.basename(relative_path) or "root"

    out_dir = os.path.join(output_folder, relative_path)
    os.makedirs(out_dir, exist_ok=True)
    excel_path = os.path.join(out_dir, f"{folder_name}.xlsx")

    if not merged_df.empty:
        # Save DataFrame
        merged_df.to_excel(excel_path, sheet_name="Part Number Details", index=False)

        # Bold "Alteration" cells
        wb = load_workbook(excel_path)
        ws = wb.active
        for r in ws.iter_rows():
            for cell in r:
                if cell.value == "Alteration":
                    cell.font = Font(bold=True)
        wb.save(excel_path)

    # Embed product images if available
    images_folder = os.path.join(root, "Product_Specification")
    if os.path.exists(images_folder):
        valid_exts = (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tiff", ".webp")
        image_files = sorted(
            f for f in glob.glob(os.path.join(images_folder, "*.*"))
            if f.lower().endswith(valid_exts)
        )
        if image_files:
            wb = load_workbook(excel_path)
            ws = wb.create_sheet(title="Images")
            for i, img_path in enumerate(image_files, start=1):
                try:
                    img = XLImage(img_path)
                    img.width = 150
                    img.height = 100
                    ws.add_image(img, f"A{i}")
                except Exception as e:
                    print(f"⚠️ Could not embed image {img_path}: {e}")
            wb.save(excel_path)

    # Process spec.html if present
    spec_file = os.path.join(root, "Product_Specification", "spec.html")
    if os.path.exists(spec_file):
        tables_with_gaps = []
        tables = extract_all_tables(spec_file)
        for table in tables:
            tables_with_gaps.append(table)
            empty_row = pd.DataFrame([[""] * table.shape[1]], columns=table.columns)
            tables_with_gaps.append(empty_row)

        parts_df = pd.concat(tables_with_gaps, ignore_index=True)
        with pd.ExcelWriter(excel_path, mode="a", engine="openpyxl") as writer:
            parts_df.to_excel(writer, sheet_name="Product_Specification", index=False)

        # Add borders
        wb = load_workbook(excel_path)
        ws = wb["Product_Specification"]
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        current_row = 2

        for table in tables:
            n_rows = table.shape[0]
            n_cols = table.shape[1]
            for r in range(current_row, current_row + n_rows):
                for c in range(1, n_cols + 1):
                    ws.cell(row=r, column=c).border = border
            current_row += n_rows + 1
        wb.save(excel_path)

    return f"✅ Saved Excel file: {excel_path}"
